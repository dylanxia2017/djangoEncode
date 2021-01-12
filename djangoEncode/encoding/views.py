from django.shortcuts import render, redirect, HttpResponse
from .forms import UserForm, RegisterForm
from . import models
import hashlib
import rarfile
import os
from io import BytesIO
import xlwt
import pandas as pd

# Create your views here.

def hash_code(s, salt='mysite'):# 加点盐
    h = hashlib.sha256()
    s += salt
    h.update(s.encode())  # update方法只接收bytes类型
    return h.hexdigest()



def upload_file(request):
    if request.method == "POST":
        File = request.FILES.get("file", None)
        if File is None:
            return HttpResponse("请选择需要上传的文件")
        else:
            # print("文件名是: ", File.name.split(".rar")[0])

            if os.path.exists(File.name.split(".rar")[0]):
                os.rmdir(File.name.split(".rar")[0])

            os.mkdir(File.name.split(".rar")[0])
            unzip_file_path = os.path.realpath(File.name.split(".rar")[0])
            # print(unzip_file_path)

            rf = rarfile.RarFile(File, mode='r')
            rf_list = rf.namelist()
            for f in rf_list:
                rf.extract(f, unzip_file_path)

            f = open("./{}/main.py".format(File.name.split(".rar")[0]), "a")

            process_str = '''
# coding=gbk
import openpyxl as xl
import pandas as pd
import os
import datetime
import csv


now_date=int(str(datetime.datetime.now().year)+str(datetime.datetime.now().month)+str(datetime.datetime.now().day))
pd.set_option('display.max_rows', 99999)
pd.set_option('display.max_columns', 99999)
files=[]


def DirAll(pathName):
    if os.path.exists(pathName):
        fileList = os.listdir(pathName);
        for f in fileList:
            if f == "$RECYCLE.BIN" or f == "System Volume Information":
                continue;
            f = os.path.join(pathName, f);
            if os.path.isdir(f):
                DirAll(f);
            else:
                if os.path.splitext(f)[1] !='.csv':
                    continue;
                else:
                    dirName = os.path.dirname(f);
                    baseName = os.path.basename(f);
                    if dirName.endswith(os.sep):
                        files.append(dirName + baseName);
                    else:
                        files.append(dirName + os.sep + baseName);

DirAll('.')

for file in files:
    path=os.path.dirname(file)
    filename = os.path.basename(file)
    # print(path)
    # print(filename)
    print(file)
    with open(file,encoding='utf8') as f:
        # for line in f.readlines():
        #     line = line.replace('  ','')
        #     # print(line)
        #     with open('.'+filename,'a',encoding='utf8') as f2:
        #         f2.write(line)
        with open('.' + filename, 'a', encoding='utf8') as f2:
            f2.write(f.read().replace('`  ','`'))

class sap_process_demo():
    def __init__(self,path):
        wb=xl.Workbook()
        wb.save('./SAP_PROCESS_RESULT.xlsx')
        df_acct = pd.read_csv(path + '/.ACTT_config_settings.csv', quoting=3, low_memory=False, delimiter='`',error_bad_lines=False)
        self.extract_date=df_acct['SettingValue NVARCHAR(1000)'][df_acct['SettingName VARCHAR(100)']=='Extract Date'].iloc[0]
        self.writer = pd.ExcelWriter('./SAP_PROCESS_RESULT.xlsx')
        self.version=df_acct['SettingValue NVARCHAR(1000)'][df_acct['SettingName VARCHAR(100)']=='SAP ERP Version'].iloc[0]
        self.writer = pd.ExcelWriter('./SAP_PROCESS_RESULT.xlsx')
        # print(self.version)

        # return

    def sap18(self, path):
        df_UST04 = pd.read_csv(path + '/.UST04.CSV',quoting=3,low_memory=False,delimiter='`')
        df_UST04_SAP_ALL_SAP_NEW = df_UST04[
            (df_UST04['PROFILE [C(00024)]'] == 'SAP_ALL') | (df_UST04['PROFILE [C(00024)]'] == 'SAP_NEW')]
        df_USR02 = pd.read_csv(path + '/.USR02.CSV',quoting=3,low_memory=False,delimiter='`')
        df_UST04_SAP_ALL_SAP_NEW_USR02 = pd.merge(df_UST04_SAP_ALL_SAP_NEW, df_USR02, how='left',
                                                  left_on='BNAME [C(00024)]', right_on='BNAME [C(00024)]')
        df_sap18 = df_UST04_SAP_ALL_SAP_NEW_USR02
        df_usr21 = pd.read_csv(path + '/.USR21.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_debug_persno = pd.merge(df_sap18, df_usr21, how='left', left_on=['BNAME [C(00024)]'],
                                         right_on=['bname [C(00024)]'])
        df_adrp = pd.read_csv(path + '/.adrp.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_persno_adrp = pd.merge(df_usr02_debug_persno, df_adrp, how='left', left_on=['persnumber [C(00020)]'],
                                        right_on=['persnumber [C(00020)]'])
        df_usr02_persno_adrp = df_usr02_persno_adrp[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                                     'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]_x',
                                                     'BNAME [C(00024)]', 'name_first [C(00080)]',
                                                     'name_last [C(00080)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr02_persno_adrp.to_excel(excel_writer=self.writer, sheet_name='sap18', index=False,
                                      header=['用户组', '账号类型', '账号有效期至', '有效期自', '锁定状态', '客户名称', '用户', '名', '姓', '最后登录日期',
                                              '最后登录时间'])
        self.writer.save()
        # print(df_sap18)

    def sap26_role(self, path):
        df_agr_1251=pd.read_csv(path + '/.AGR_1251.CSV',quoting=3,low_memory=False,delimiter='`')
        df_agr_1251['LOW [C(00080)]'] =df_agr_1251['LOW [C(00080)]'].astype(str)
        df_agr_1251_ACTVT_02 = df_agr_1251[
            (df_agr_1251['OBJCT [C(00020)]'] == 'S_DEVELOP')&(df_agr_1251['FIELD [C(00020)]'] == 'ACTVT')&(df_agr_1251['LOW [C(00080)]'] =='02')].drop_duplicates(subset='AGR_NAME [C(00060)]',keep='first',inplace=False)['AGR_NAME [C(00060)]']
        # print(df_agr_1251_ACTVT_02)
        df_agr_1251_ACTVT_03 = df_agr_1251[
           (df_agr_1251['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                        df_agr_1251['FIELD [C(00020)]'] == 'ACTVT') & (
                        df_agr_1251['LOW [C(00080)]'] == '03')].drop_duplicates(subset='AGR_NAME [C(00060)]',keep='first', inplace=False)['AGR_NAME [C(00060)]']
        df_agr_1251_OBJTYPE_DEBUG = df_agr_1251[
            (df_agr_1251['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                    df_agr_1251['FIELD [C(00020)]'] == 'OBJTYPE') & (
                    df_agr_1251['LOW [C(00080)]'] == 'DEBUG')].drop_duplicates(subset='AGR_NAME [C(00060)]', keep='first',
                                                                            inplace=False)['AGR_NAME [C(00060)]']
        df_agr_1251_ACTVT_all = df_agr_1251[
            (df_agr_1251['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                    df_agr_1251['FIELD [C(00020)]'] == 'ACTVT') & (
                    df_agr_1251['LOW [C(00080)]'] == '*')].drop_duplicates(subset='AGR_NAME [C(00060)]', keep='first',
                                                                            inplace=False)['AGR_NAME [C(00060)]']
        df_agr_1251_OBJTYPE_all = df_agr_1251[
            (df_agr_1251['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                    df_agr_1251['FIELD [C(00020)]'] == 'OBJTYPE') & (
                    df_agr_1251['LOW [C(00080)]'] == '*')].drop_duplicates(subset='AGR_NAME [C(00060)]', keep='first',
                                                                           inplace=False)['AGR_NAME [C(00060)]']
        df_agr_name_23debug=pd.merge(df_agr_1251_ACTVT_02,df_agr_1251_ACTVT_03,on=['AGR_NAME [C(00060)]'])
        df_agr_name_23debug=pd.merge(df_agr_1251_OBJTYPE_DEBUG,df_agr_name_23debug,on=['AGR_NAME [C(00060)]'])
        df_agr_name_23all=pd.merge(df_agr_1251_ACTVT_02,df_agr_1251_ACTVT_03,on=['AGR_NAME [C(00060)]'])
        df_agr_name_23all=pd.merge(df_agr_name_23all,df_agr_1251_OBJTYPE_all,on=['AGR_NAME [C(00060)]'])
        df_agr_name_alldebug=pd.merge(df_agr_1251_ACTVT_all,df_agr_1251_OBJTYPE_DEBUG,on=['AGR_NAME [C(00060)]'])
        df_agr_name_allall=pd.merge(df_agr_1251_ACTVT_all,df_agr_1251_OBJTYPE_all,on=['AGR_NAME [C(00060)]'])
        df_agr_name=pd.concat([df_agr_name_23debug, df_agr_name_23all,df_agr_name_alldebug,df_agr_name_allall]).drop_duplicates()
        # print(df_agr_name)

    def sap26(self, path):
        #通过ust12筛选拥有debug的权限
        df_ust12=pd.read_csv(path + '/.UST12.CSV',quoting=3,low_memory=False,delimiter='`')
        # df_ust12['VON [C(00080)]']=df_ust12['VON [C(00080)]'].astype(str)
        df_ust12_ACTVT_02=df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_DEVELOP')&(df_ust12['FIELD [C(00020)]'] == 'ACTVT')&(df_ust12['VON [C(00080)]'] =='02')].drop_duplicates(subset='AUTH [C(00024)]',keep='first',inplace=False)['AUTH [C(00024)]']
        # print(df_agr_1251_ACTVT_02)
        df_ust12_ACTVT_03 = df_ust12[
           (df_ust12['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                        df_ust12['FIELD [C(00020)]'] == 'ACTVT') & (
                        df_ust12['VON [C(00080)]'] == '03')].drop_duplicates(subset='AUTH [C(00024)]',keep='first', inplace=False)['AUTH [C(00024)]']
        df_ust12_OBJTYPE_DEBUG = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                    df_ust12['FIELD [C(00020)]'] == 'OBJTYPE') & (
                    df_ust12['VON [C(00080)]'] == 'DEBUG')].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                            inplace=False)['AUTH [C(00024)]']
        df_ust12_ACTVT_all = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                    df_ust12['FIELD [C(00020)]'] == 'ACTVT') & (
                    df_ust12['VON [C(00080)]'] == '*')].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                            inplace=False)['AUTH [C(00024)]']
        df_ust12_OBJTYPE_all = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_DEVELOP') & (
                    df_ust12['FIELD [C(00020)]'] == 'OBJTYPE') & (
                    df_ust12['VON [C(00080)]'] == '*')].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                           inplace=False)['AUTH [C(00024)]']
        df_auth_23debug=pd.merge(df_ust12_ACTVT_02,df_ust12_ACTVT_03,on=['AUTH [C(00024)]'])
        df_auth_23debug=pd.merge(df_ust12_OBJTYPE_DEBUG,df_auth_23debug,on=['AUTH [C(00024)]'])
        df_auth_23all=pd.merge(df_ust12_ACTVT_02,df_ust12_ACTVT_03,on=['AUTH [C(00024)]'])
        df_auth_23all=pd.merge(df_auth_23all,df_ust12_OBJTYPE_all,on=['AUTH [C(00024)]'])
        df_auth_alldebug=pd.merge(df_ust12_ACTVT_all,df_ust12_OBJTYPE_DEBUG,on=['AUTH [C(00024)]'])
        df_auth_allall=pd.merge(df_ust12_ACTVT_all,df_ust12_OBJTYPE_all,on=['AUTH [C(00024)]'])
        df_auth=pd.concat([df_auth_23debug, df_auth_23all,df_auth_alldebug,df_auth_allall]).drop_duplicates()
        df_auth.columns=['AUTH']
        # print(df_auth)
        # 通过ust10s将auth归于profn
        df_ust10s=pd.read_csv(path + '/.UST10S.CSV',quoting=3,low_memory=False,delimiter='`')
        df_auth_profn=pd.merge(df_ust10s,df_auth,how='left',left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn=df_auth_profn[df_auth_profn['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')['PROFN [C(00024)]']
        # print(df_auth_profn)
        #通过ust10c将部分profn归于subprof
        df_ust10c=pd.read_csv(path + '/.UST10C.CSV',quoting=3,low_memory=False,delimiter='`')
        df_profn_subprof = pd.merge(df_ust10c, df_auth_profn, how='left', right_on=['PROFN [C(00024)]'],left_on='SUBPROF [C(00024)]')
        df_profn_subprof=df_profn_subprof[df_profn_subprof['PROFN [C(00024)]_y'].notna()]['PROFN [C(00024)]_x'].drop_duplicates()
        #获得profn和subprof的并集
        df_prof = pd.concat([df_profn_subprof,df_auth_profn]).drop_duplicates()
        df_prof.name='PROFN [C(00024)]'
        #通过ust04获得prof对应userid
        df_ust04=pd.read_csv(path + '/.UST04.CSV',quoting=3,low_memory=False,delimiter='`')
        df_ust04_prof = pd.merge(df_ust04, df_prof, how='left', left_on=['PROFILE [C(00024)]'],right_on='PROFN [C(00024)]')
        df_ust04_prof=df_ust04_prof[df_ust04_prof['PROFN [C(00024)]'].notna()]['BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_usr02=pd.read_csv(path + '/.USR02.CSV',quoting=3,low_memory=False,delimiter='`')
        df_ust04_prof.name='BNAME'
        df_usr02_debug = pd.merge(df_usr02, df_ust04_prof, how='left', left_on=['BNAME [C(00024)]'],right_on='BNAME')
        df_usr02_debug=df_usr02_debug[df_usr02_debug['BNAME'].notna()]
        df_usr02_debug = df_usr02_debug[(df_usr02_debug['USTYP [C(00002)]']=='A')&(df_usr02_debug['UFLAG [b(00001)]']!=64)&((df_usr02_debug['GLTGB [D(00016)]']==0)|(df_usr02_debug['GLTGB [D(00016)]']>=int(self.extract_date)))]
        df_usr02_debug=df_usr02_debug[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
               'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
               'BNAME [C(00024)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr21=pd.read_csv(path + '/.USR21.CSV',quoting=3,low_memory=False,delimiter='`')
        df_usr02_debug_persno=pd.merge(df_usr02_debug, df_usr21, how='left', left_on=['BNAME [C(00024)]'],right_on=['bname [C(00024)]'])
        df_adrp=pd.read_csv(path + '/.adrp.CSV',quoting=3,low_memory=False,delimiter='`')
        df_usr02_persno_adrp=pd.merge(df_usr02_debug_persno, df_adrp, how='left', left_on=['persnumber [C(00020)]'],right_on=['persnumber [C(00020)]'])
        df_usr02_persno_adrp=df_usr02_persno_adrp[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
               'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]','BNAME [C(00024)]','name_first [C(00080)]','name_last [C(00080)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr02_persno_adrp.to_excel(excel_writer=self.writer, sheet_name='sap26', index=False,
                          header=['用户组', '账号类型', '账号有效期至', '有效期自', '锁定状态','客户名称','用户','名','姓','最后登录日期', '最后登录时间'])
        self.writer.save()
        # print(df_usr02_persno_adrp.columns)

    def sap05(self,path):
        # 导入sqpwlserv，获取服务器名称
        df_SAPWLSERV = pd.read_csv(path + '/.SAPWLSERV.CSV', quoting=3, low_memory=False, delimiter='`')
        df_SAPWLSERV_hostlist=df_SAPWLSERV.drop_duplicates(subset='HOST [C(00064)]')['HOST [C(00064)]'].tolist()
        # 导入PAHI,筛选PARSTATE [C(00002)]为‘A’，hostname在sqpwlserv里的
        df_PAHI=pd.read_csv(path + '/.PAHI.CSV', quoting=3, low_memory=False, delimiter='`')
        df_PAHI=df_PAHI[df_PAHI['PARSTATE [C(00002)]']=='A']
        df_PAHI = pd.merge(df_PAHI, df_SAPWLSERV['HOST [C(00064)]'], how='left', left_on=['HOSTNAME [C(00064)]'], right_on='HOST [C(00064)]')
        df_PAHI = df_PAHI[df_PAHI['HOST [C(00064)]'].notna()]
        #获取密码策略
        # accept_sso2_ticket=df_PAHI[df_PAHI['PARNAME [C(00128)]']=='login/accept_sso2_ticket']['PARVALUE [C(00128)]'].iloc[0]
        # create_sso2_ticket =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/create_sso2_ticket']['PARVALUE [C(00128)]'].iloc[0]
        failed_user_auto_unlock =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/failed_user_auto_unlock'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        fails_to_user_lock =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/fails_to_user_lock'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        min_password_diff =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/min_password_diff'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        min_password_digits =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/min_password_digits'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        min_password_letters =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/min_password_letters'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        min_password_lng =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/min_password_lng'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        min_password_specials =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/min_password_specials'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        password_expiration_time =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/password_expiration_time'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        password_history_size =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/password_history_size'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        gui_auto_logout =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'rdisp/gui_auto_logout'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        enable =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'rsau/enable'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]
        no_automatic_user_sapstar =df_PAHI[df_PAHI['PARNAME [C(00128)]'] == 'login/no_automatic_user_sapstar'][['PARVALUE [C(00128)]','HOSTNAME [C(00064)]']]


        # print(df_SAPWLSERV_hostlist)

        result_wb=xl.load_workbook('./SAP_PROCESS_RESULT.xlsx')
        result_wb.create_sheet('sap05')
        ws_sap05=result_wb['sap05']
        for host in df_SAPWLSERV_hostlist:
            ws_sap05.append(['参数名称','行业标准','服务器：'+host])
            try:
                ws_sap05.append(['failed_user_auto_unlock','0',failed_user_auto_unlock[failed_user_auto_unlock['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['fails_to_user_lock', '6', fails_to_user_lock[fails_to_user_lock['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['min_password_diff', '1', min_password_diff[min_password_diff['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['min_password_digits', '1', min_password_digits[min_password_digits['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['min_password_letters', '1', min_password_letters[min_password_letters['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['min_password_lng', '8', min_password_lng[min_password_lng['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['min_password_specials', '1', min_password_specials[min_password_specials['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['password_expiration_time', '90', password_expiration_time[password_expiration_time['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['password_history_size', '12', password_history_size[password_history_size['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['gui_auto_logout', '1800', gui_auto_logout[gui_auto_logout['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['SM20_enable', '1', enable[enable['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['no_automatic_user_sapstar', '1', no_automatic_user_sapstar[no_automatic_user_sapstar['HOSTNAME [C(00064)]']==host]['PARVALUE [C(00128)]'].iloc[0]])
                ws_sap05.append(['---------------------------------------------------', '  ', ' '])
                ws_sap05.append(['  ', '  ', ' '])
                result_wb.save('./SAP_PROCESS_RESULT.xlsx')
            except:
                pass
        # print( df_SAPWLSERV_hostlist)

    def sap06(self,path):
        # 通过ust12筛选拥有debug的权限
        df_ust12 = pd.read_csv(path + '/.UST12.CSV', quoting=3, low_memory=False, delimiter='`')
        # df_ust12['VON [C(00080)]']=df_ust12['VON [C(00080)]'].astype(str)
        df_ust12_PFCG = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_TCODE') & (df_ust12['FIELD [C(00020)]'] == 'TCD') & ((df_ust12['VON [C(00080)]'] == '*')|(df_ust12['VON [C(00080)]'] == 'PFCG'))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                             inplace=False)['AUTH [C(00024)]']

        # print(df_ust12_PFCG)
        df_ust12_USR_AGR_01 = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_AGR') & (df_ust12['FIELD [C(00020)]'] == 'ACTVT') & (df_ust12['VON [C(00080)]'] == '01')].drop_duplicates(subset='AUTH [C(00024)]', keep='first',inplace=False)['AUTH [C(00024)]']
        df_ust12_USR_AGR_02 = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_AGR') & (df_ust12['FIELD [C(00020)]'] == 'ACTVT') & (df_ust12['VON [C(00080)]'] == '02')].drop_duplicates(
            subset='AUTH [C(00024)]', keep='first', inplace=False)['AUTH [C(00024)]']
        df_ust12_USR_AGR_64 = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_AGR')& (df_ust12['FIELD [C(00020)]'] == 'ACTVT')  & (df_ust12['VON [C(00080)]'] == '64')].drop_duplicates(
            subset='AUTH [C(00024)]', keep='first', inplace=False)['AUTH [C(00024)]']
        df_ust12_USR_AGR_ALL = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_AGR')& (df_ust12['FIELD [C(00020)]'] == 'ACTVT')  & (df_ust12['VON [C(00080)]'] == '*')].drop_duplicates(
            subset='AUTH [C(00024)]', keep='first', inplace=False)['AUTH [C(00024)]']
        df_ust12_USR_AGR=pd.merge(df_ust12_USR_AGR_01, df_ust12_USR_AGR_02, on=['AUTH [C(00024)]'])
        df_ust12_USR_AGR=pd.merge(df_ust12_USR_AGR, df_ust12_USR_AGR_64, on=['AUTH [C(00024)]'])
        df_ust12_USR_AGR=pd.concat([df_ust12_USR_AGR['AUTH [C(00024)]'], df_ust12_USR_AGR_ALL]).drop_duplicates()


        df_ust12_USER_PRO = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_PRO') & ((
                    df_ust12['VON [C(00080)]'] == '01')|(df_ust12['VON [C(00080)]'] == '*'))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                            inplace=False)['AUTH [C(00024)]']
        df_ust12_USER_TCD_USER_VAL  = df_ust12[
            ((df_ust12['OBJCT [C(00020)]'] == 'S_USER_TCD') & (
                    df_ust12['VON [C(00080)]'] == '*'))|((df_ust12['OBJCT [C(00020)]'] == 'S_USER_VAL') & (df_ust12['FIELD [C(00020)]'] == 'AUTH_VALUE') & (
                    df_ust12['VON [C(00080)]'] == '*'))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                        inplace=False)['AUTH [C(00024)]']

        # df_auth_all = pd.merge(df_ust12_PFCG, df_ust12_USR_AGR, on=['AUTH [C(00024)]'])
        # df_auth_all = pd.merge(df_auth_all, df_ust12_USER_PRO, on=['AUTH [C(00024)]'])
        # df_auth_all = pd.merge(df_auth_all, df_ust12_USER_TCD_USER_VAL, on=['AUTH [C(00024)]'])['AUTH [C(00024)]']
        # df_auth_all=pd.concat([df_ust12_PFCG, df_ust12_USR_AGR,df_ust12_USER_PRO,df_ust12_USER_TCD_USER_VAL]).drop_duplicates()

        ##PCFG
        # 通过ust10s将auth归于profn
        df_ust10s = pd.read_csv(path + '/.UST10S.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust12_PFCG.name='AUTH'
        # print(df_ust12_PFCG)
        df_auth_profn_PFCG = pd.merge(df_ust10s,df_ust12_PFCG,  how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_PFCG =df_auth_profn_PFCG[df_auth_profn_PFCG['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']
        # print(df_auth_profn_PFCG)

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_PCFG = pd.merge(df_ust10c,df_auth_profn_PFCG,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        # print(df_profn_subprof_PCFG.columns)
        df_profn_subprof_PCFG = df_profn_subprof_PCFG[df_profn_subprof_PCFG['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_PCFG = pd.concat([df_profn_subprof_PCFG, df_auth_profn_PFCG]).drop_duplicates()
        df_prof_PCFG.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_PCFG = pd.merge(df_ust04, df_prof_PCFG, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_PCFG = df_ust04_prof_PCFG[df_ust04_prof_PCFG['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_PCFG.name = 'BNAME'
        df_usr02_PCFG = pd.merge(df_usr02, df_ust04_prof_PCFG, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_PCFG = df_usr02_PCFG[df_usr02_PCFG['BNAME'].notna()]


        ##df_ust12_USR_AGR
        # 通过ust10s将auth归于profn
        df_ust12_USR_AGR.name='AUTH'
        df_auth_profn_USR_AGR = pd.merge(df_ust10s, df_ust12_USR_AGR, how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_USR_AGR =df_auth_profn_USR_AGR[df_auth_profn_USR_AGR['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_USR_AGR = pd.merge(df_ust10c,df_auth_profn_USR_AGR,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        df_profn_subprof_USR_AGR = df_profn_subprof_USR_AGR[df_profn_subprof_USR_AGR['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_USR_AGR = pd.concat([df_profn_subprof_USR_AGR, df_auth_profn_USR_AGR]).drop_duplicates()
        df_prof_USR_AGR.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_USR_AGR = pd.merge(df_ust04, df_prof_USR_AGR, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_USR_AGR = df_ust04_prof_USR_AGR[df_ust04_prof_USR_AGR['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_USR_AGR.name = 'BNAME'
        df_usr02_USR_AGR = pd.merge(df_usr02, df_ust04_prof_USR_AGR, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_USR_AGR = df_usr02_USR_AGR[df_usr02_USR_AGR['BNAME'].notna()]


        ##df_ust12_USER_PRO
        # 通过ust10s将auth归于profn
        df_ust12_USER_PRO.name='AUTH'
        df_auth_profn_USER_PRO = pd.merge(df_ust10s,df_ust12_USER_PRO,  how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_USER_PRO =df_auth_profn_USER_PRO[df_auth_profn_USER_PRO['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']
        # 通过ust10c将部分profn归于subprof
        df_profn_subprof_USER_PRO = pd.merge(df_auth_profn_USER_PRO, df_ust10c, how='left', left_on=['PROFN [C(00024)]'],
                                    right_on='SUBPROF [C(00024)]')
        df_profn_subprof_USER_PRO = df_profn_subprof_USER_PRO[df_profn_subprof_USER_PRO['SUBPROF [C(00024)]'].notna()][
            'PROFN [C(00024)]_y'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_USER_PRO = pd.concat([df_profn_subprof_USER_PRO, df_auth_profn_USER_PRO]).drop_duplicates()
        df_prof_USER_PRO.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04_prof_USER_PRO = pd.merge(df_ust04, df_prof_USER_PRO, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_USER_PRO = df_ust04_prof_USER_PRO[df_ust04_prof_USER_PRO['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_USER_PRO.name = 'BNAME'
        df_usr02_USER_PRO = pd.merge(df_usr02, df_ust04_prof_USER_PRO, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_USER_PRO = df_usr02_USER_PRO[df_usr02_USER_PRO['BNAME'].notna()]

        ##df_ust12_USER_TCD_USER_VAL
        # 通过ust10s将auth归于profn
        df_ust12_USER_TCD_USER_VAL.name='AUTH'
        df_auth_profn_USER_TCD_USER_VAL = pd.merge(df_ust10s,df_ust12_USER_TCD_USER_VAL,  how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_USER_TCD_USER_VAL =df_auth_profn_USER_TCD_USER_VAL[df_auth_profn_USER_TCD_USER_VAL['PROFN [C(00024)]'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']

        # 通过ust10c将部分profn归于subprof
        df_profn_subprof_USER_TCD_USER_VAL = pd.merge(df_auth_profn_USER_TCD_USER_VAL, df_ust10c, how='left', left_on=['PROFN [C(00024)]'],
                                    right_on='SUBPROF [C(00024)]')
        df_profn_subprof_USER_TCD_USER_VAL = df_profn_subprof_USER_TCD_USER_VAL[df_profn_subprof_USER_TCD_USER_VAL['SUBPROF [C(00024)]'].notna()][
            'PROFN [C(00024)]_y'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_USER_TCD_USER_VAL = pd.concat([df_profn_subprof_USER_TCD_USER_VAL, df_auth_profn_USER_TCD_USER_VAL]).drop_duplicates()
        df_prof_USER_TCD_USER_VAL.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04_prof_USER_TCD_USER_VAL = pd.merge(df_ust04, df_prof_USER_TCD_USER_VAL, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_USER_TCD_USER_VAL = df_ust04_prof_USER_TCD_USER_VAL[df_ust04_prof_USER_TCD_USER_VAL['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_ust04_prof_USER_TCD_USER_VAL.name = 'BNAME'
        df_usr02_USER_TCD_USER_VAL = pd.merge(df_usr02, df_ust04_prof_USER_TCD_USER_VAL, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_USER_TCD_USER_VAL = df_usr02_USER_TCD_USER_VAL[df_usr02_USER_TCD_USER_VAL['BNAME'].notna()]

        df_usr02_a=pd.merge(df_usr02_PCFG, df_usr02_USR_AGR['BNAME [C(00024)]'], on=['BNAME [C(00024)]'])
        df_usr02_a=pd.merge(df_usr02_a, df_usr02_USER_PRO['BNAME [C(00024)]'], on=['BNAME [C(00024)]'])
        df_usr02_a = pd.merge(df_usr02_a, df_usr02_USER_TCD_USER_VAL['BNAME [C(00024)]'], on=['BNAME [C(00024)]'])

        # print(df_usr02_PCFG.shape)

        df_usr02_a = df_usr02_a[
            (df_usr02_a['USTYP [C(00002)]'] == 'A') & (df_usr02_a['UFLAG [b(00001)]'] != 64) & (
                        (df_usr02_a['GLTGB [D(00016)]'] == 0) | (
                            df_usr02_a['GLTGB [D(00016)]'] >= int(self.extract_date)))]
        df_usr02_debug = df_usr02_a[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                         'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
                                         'BNAME [C(00024)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr21 = pd.read_csv(path + '/.USR21.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_a_persno = pd.merge(df_usr02_debug, df_usr21, how='left', left_on=['BNAME [C(00024)]'],
                                         right_on=['bname [C(00024)]'])
        df_adrp = pd.read_csv(path + '/.adrp.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_persno_adrp = pd.merge(df_usr02_a_persno, df_adrp, how='left', left_on=['persnumber [C(00020)]'],
                                        right_on=['persnumber [C(00020)]'])
        df_usr02_persno_adrp = df_usr02_persno_adrp[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                                     'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
                                                     'BNAME [C(00024)]', 'name_first [C(00080)]',
                                                     'name_last [C(00080)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr02_persno_adrp.to_excel(excel_writer=self.writer, sheet_name='sap06', index=False,
                                      header=['用户组', '账号类型', '账号有效期至', '有效期自', '锁定状态', '客户名称', '用户', '名', '姓', '最后登录日期',
                                              '最后登录时间'])
        self.writer.save()
        # print( df_prof)
        # print(df_usr02_PCFG.shape)


    def sap06b(self,path):
        # 通过ust12筛选拥有debug的权限
        df_ust12 = pd.read_csv(path + '/.UST12.CSV', quoting=3, low_memory=False, delimiter='`')
        # df_ust12['VON [C(00080)]']=df_ust12['VON [C(00080)]'].astype(str)
        df_ust12_S_TCODE = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_TCODE') & (df_ust12['FIELD [C(00020)]'] == 'TCD') & ((df_ust12['VON [C(00080)]'] == '*')|(df_ust12['VON [C(00080)]'] == 'SU01')|(df_ust12['VON [C(00080)]'] == 'SU01_NAV')
                                                                                                     |(df_ust12['VON [C(00080)]'] == 'SU10')|(df_ust12['VON [C(00080)]'] == 'SU12')|(df_ust12['VON [C(00080)]'] == 'OY27')|(df_ust12['VON [C(00080)]'] == 'OY30')|(df_ust12['VON [C(00080)]'] == 'OY28')|(df_ust12['VON [C(00080)]'] == 'OY29')|(df_ust12['VON [C(00080)]'] == 'OOUS')|(df_ust12['VON [C(00080)]'] == 'OTZ1')|(df_ust12['VON [C(00080)]'] == 'OMDL')|(df_ust12['VON [C(00080)]'] == 'OMEH')
                                                                                                     |(df_ust12['VON [C(00080)]'] == 'OMWF')|(df_ust12['VON [C(00080)]'] == 'OPF0')|(df_ust12['VON [C(00080)]'] == 'GCE1'))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                             inplace=False)['AUTH [C(00024)]']


        df_ust12_S_USER_GRP= df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_GRP') & (df_ust12['FIELD [C(00020)]'] == 'ACTVT') & ((
                    df_ust12['VON [C(00080)]'] == '01')|(df_ust12['VON [C(00080)]'] == '*')|(df_ust12['VON [C(00080)]'] == '02'))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                            inplace=False)['AUTH [C(00024)]']

        # df_auth_all = pd.merge(df_ust12_PFCG, df_ust12_USR_AGR, on=['AUTH [C(00024)]'])
        # df_auth_all = pd.merge(df_auth_all, df_ust12_USER_PRO, on=['AUTH [C(00024)]'])
        # df_auth_all = pd.merge(df_auth_all, df_ust12_USER_TCD_USER_VAL, on=['AUTH [C(00024)]'])['AUTH [C(00024)]']
        # df_auth_all=pd.concat([df_ust12_PFCG, df_ust12_USR_AGR,df_ust12_USER_PRO,df_ust12_USER_TCD_USER_VAL]).drop_duplicates()

        ##df_ust12_S_TCODE
        # 通过ust10s将auth归于profn
        df_ust10s = pd.read_csv(path + '/.UST10S.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust12_S_TCODE.name='AUTH'
        # print(df_ust12_S_TCODE)
        df_auth_profn_S_TCODE = pd.merge(df_ust10s,df_ust12_S_TCODE,  how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_S_TCODE =df_auth_profn_S_TCODE[df_auth_profn_S_TCODE['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']
        # print(df_auth_profn_PFCG)

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_S_TCODE = pd.merge(df_ust10c,df_auth_profn_S_TCODE,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        # print(df_profn_subprof_PCFG.columns)
        df_profn_subprof_S_TCODE = df_profn_subprof_S_TCODE[df_profn_subprof_S_TCODE['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_S_TCODE = pd.concat([df_profn_subprof_S_TCODE, df_auth_profn_S_TCODE]).drop_duplicates()
        df_prof_S_TCODE.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_TCODE = pd.merge(df_ust04, df_prof_S_TCODE, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_S_TCODE = df_ust04_prof_S_TCODE[df_ust04_prof_S_TCODE['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_TCODE.name = 'BNAME'
        df_usr02_S_TCODE = pd.merge(df_usr02, df_ust04_prof_S_TCODE, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_S_TCODE = df_usr02_S_TCODE[df_usr02_S_TCODE['BNAME'].notna()]


        ##df_ust12_S_USER_GRP
        # 通过ust10s将auth归于profn
        df_ust12_S_USER_GRP.name='AUTH'
        df_auth_profn_USER_GRP = pd.merge(df_ust10s, df_ust12_S_USER_GRP, how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_USER_GRP =df_auth_profn_USER_GRP[df_auth_profn_USER_GRP['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_USER_GRP = pd.merge(df_ust10c,df_auth_profn_USER_GRP,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        df_profn_subprof_USER_GRP = df_profn_subprof_USER_GRP[df_profn_subprof_USER_GRP['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_USER_GRP = pd.concat([df_profn_subprof_USER_GRP, df_auth_profn_USER_GRP]).drop_duplicates()
        df_prof_USER_GRP.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_USER_GRP = pd.merge(df_ust04, df_prof_USER_GRP, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_USER_GRP = df_ust04_prof_USER_GRP[df_ust04_prof_USER_GRP['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_USER_GRP.name = 'BNAME'
        df_usr02_USER_GRP = pd.merge(df_usr02, df_ust04_prof_USER_GRP, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_USER_GRP = df_usr02_USER_GRP[df_usr02_USER_GRP['BNAME'].notna()]


        df_usr02_a=pd.merge(df_usr02_S_TCODE, df_usr02_USER_GRP['BNAME [C(00024)]'], on=['BNAME [C(00024)]'])

        # print(df_usr02_PCFG.shape)

        df_usr02_a = df_usr02_a[
            (df_usr02_a['USTYP [C(00002)]'] == 'A') & (df_usr02_a['UFLAG [b(00001)]'] != 64) & (
                        (df_usr02_a['GLTGB [D(00016)]'] == 0) | (
                            df_usr02_a['GLTGB [D(00016)]'] >= int(self.extract_date)))]
        df_usr02_debug = df_usr02_a[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                         'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
                                         'BNAME [C(00024)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr21 = pd.read_csv(path + '/.USR21.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_a_persno = pd.merge(df_usr02_debug, df_usr21, how='left', left_on=['BNAME [C(00024)]'],
                                         right_on=['bname [C(00024)]'])
        df_adrp = pd.read_csv(path + '/.adrp.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_persno_adrp = pd.merge(df_usr02_a_persno, df_adrp, how='left', left_on=['persnumber [C(00020)]'],
                                        right_on=['persnumber [C(00020)]'])
        df_usr02_persno_adrp = df_usr02_persno_adrp[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                                     'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
                                                     'BNAME [C(00024)]', 'name_first [C(00080)]',
                                                     'name_last [C(00080)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr02_persno_adrp.to_excel(excel_writer=self.writer, sheet_name='sap06b', index=False,
                                      header=['用户组', '账号类型', '账号有效期至', '有效期自', '锁定状态', '客户名称', '用户', '名', '姓', '最后登录日期',
                                              '最后登录时间'])
        self.writer.save()
        # print( df_prof)
        # print(df_usr02_PCFG.shape)


    def sap13(self,path):
        # 通过ust12筛选拥有debug的权限
        df_ust12 = pd.read_csv(path + '/.UST12.CSV', quoting=3, low_memory=False, delimiter='`')
        # df_ust12['VON [C(00080)]']=df_ust12['VON [C(00080)]'].astype(str)
        df_ust12_S_TCODE = df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_TCODE') & (df_ust12['FIELD [C(00020)]'] == 'TCD') & ((df_ust12['VON [C(00080)]'] == '*')|(df_ust12['VON [C(00080)]'] == 'STMS')|(df_ust12['VON [C(00080)]'] == 'STMS_IMPORT')
                                                                                                     )].drop_duplicates(subset='AUTH [C(00024)]', keep='first',inplace=False)['AUTH [C(00024)]']

        df_ust12_S_TRANSPRT= df_ust12[
            (df_ust12['OBJCT [C(00020)]'] == 'S_USER_GRP') & (df_ust12['FIELD [C(00020)]'] == 'ACTVT') & ((
                    df_ust12['VON [C(00080)]'] == '03')|(df_ust12['VON [C(00080)]'] == '*'))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                                                                            inplace=False)['AUTH [C(00024)]']

        df_ust12_S_CTS_ADMI= df_ust12[((
            (df_ust12['OBJCT [C(00020)]'] == 'S_CTS_ADMI')  & ((
                    df_ust12['VON [C(00080)]'] == 'IMPS')|(df_ust12['VON [C(00080)]'] == 'IMPA)')|(df_ust12['VON [C(00080)]'] == '*')))|((df_ust12['OBJCT [C(00020)]'] == 'S_CTS_SADM')  & ((
                    df_ust12['VON [C(00080)]'] == 'IMPS')|(df_ust12['VON [C(00080)]'] == 'IMPA)')|(df_ust12['VON [C(00080)]'] == '*'))))].drop_duplicates(subset='AUTH [C(00024)]', keep='first',
                     inplace=False)['AUTH [C(00024)]']

        ##df_ust12_S_TCODE
        # 通过ust10s将auth归于profn
        df_ust10s = pd.read_csv(path + '/.UST10S.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust12_S_TCODE.name='AUTH'
        # print(df_ust12_S_TCODE)
        df_auth_profn_S_TCODE = pd.merge(df_ust10s,df_ust12_S_TCODE,  how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_S_TCODE =df_auth_profn_S_TCODE[df_auth_profn_S_TCODE['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']
        # print(df_auth_profn_PFCG)

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_S_TCODE = pd.merge(df_ust10c,df_auth_profn_S_TCODE,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        # print(df_profn_subprof_PCFG.columns)
        df_profn_subprof_S_TCODE = df_profn_subprof_S_TCODE[df_profn_subprof_S_TCODE['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_S_TCODE = pd.concat([df_profn_subprof_S_TCODE, df_auth_profn_S_TCODE]).drop_duplicates()
        df_prof_S_TCODE.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_TCODE = pd.merge(df_ust04, df_prof_S_TCODE, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_S_TCODE = df_ust04_prof_S_TCODE[df_ust04_prof_S_TCODE['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_TCODE.name = 'BNAME'
        df_usr02_S_TCODE = pd.merge(df_usr02, df_ust04_prof_S_TCODE, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_S_TCODE = df_usr02_S_TCODE[df_usr02_S_TCODE['BNAME'].notna()]


        ##df_ust12_S_TRANSPRT
        # 通过ust10s将auth归于profn
        df_ust12_S_TRANSPRT.name='AUTH'
        df_auth_profn_S_TRANSPRT = pd.merge(df_ust10s, df_ust12_S_TRANSPRT, how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_S_TRANSPRT =df_auth_profn_S_TRANSPRT[df_auth_profn_S_TRANSPRT['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_S_TRANSPRT = pd.merge(df_ust10c,df_auth_profn_S_TRANSPRT,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        df_profn_subprof_S_TRANSPRT = df_profn_subprof_S_TRANSPRT[df_profn_subprof_S_TRANSPRT['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_S_TRANSPRT = pd.concat([df_profn_subprof_S_TRANSPRT, df_auth_profn_S_TRANSPRT]).drop_duplicates()
        df_prof_S_TRANSPRT.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_TRANSPRT = pd.merge(df_ust04, df_prof_S_TRANSPRT, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_S_TRANSPRT = df_ust04_prof_S_TRANSPRT[df_ust04_prof_S_TRANSPRT['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_TRANSPRT.name = 'BNAME'
        df_usr02_S_TRANSPRT = pd.merge(df_usr02, df_ust04_prof_S_TRANSPRT, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_S_TRANSPRT = df_usr02_S_TRANSPRT[df_usr02_S_TRANSPRT['BNAME'].notna()]

        ##df_ust12_S_CTS_ADMI
        # 通过ust10s将auth归于profn
        df_ust12_S_CTS_ADMI.name='AUTH'
        df_auth_profn_S_CTS_ADMI = pd.merge(df_ust10s, df_ust12_S_CTS_ADMI, how='left', left_on=['AUTH [C(00024)]'],right_on=['AUTH'])
        df_auth_profn_S_CTS_ADMI =df_auth_profn_S_CTS_ADMI[df_auth_profn_S_CTS_ADMI['AUTH'].notna()].drop_duplicates(subset='PROFN [C(00024)]')[
            'PROFN [C(00024)]']

        # 通过ust10c将部分profn归于subprof
        df_ust10c = pd.read_csv(path + '/.UST10C.CSV', quoting=3, low_memory=False, delimiter='`')
        df_profn_subprof_S_CTS_ADMI = pd.merge(df_ust10c,df_auth_profn_S_CTS_ADMI,  how='left', right_on=['PROFN [C(00024)]'],
                                    left_on='SUBPROF [C(00024)]')
        df_profn_subprof_S_CTS_ADMI = df_profn_subprof_S_CTS_ADMI[df_profn_subprof_S_CTS_ADMI['PROFN [C(00024)]_y'].notna()][
            'PROFN [C(00024)]_x'].drop_duplicates()
        # 获得profn和subprof的并集
        df_prof_S_CTS_ADMI = pd.concat([df_profn_subprof_S_CTS_ADMI, df_auth_profn_S_CTS_ADMI]).drop_duplicates()
        df_prof_S_CTS_ADMI.name = 'PROFN [C(00024)]'
        # 通过ust04获得prof对应userid
        df_ust04 = pd.read_csv(path + '/.UST04.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_CTS_ADMI = pd.merge(df_ust04, df_prof_S_CTS_ADMI, how='left', left_on=['PROFILE [C(00024)]'],
                                 right_on='PROFN [C(00024)]')
        df_ust04_prof_S_CTS_ADMI = df_ust04_prof_S_CTS_ADMI[df_ust04_prof_S_CTS_ADMI['PROFN [C(00024)]'].notna()][
            'BNAME [C(00024)]'].drop_duplicates()
        # 通过usr02筛选type为A，lock不为64，有效期0或9999或大于提取日期
        df_usr02 = pd.read_csv(path + '/.USR02.CSV', quoting=3, low_memory=False, delimiter='`')
        df_ust04_prof_S_CTS_ADMI.name = 'BNAME'
        df_usr02_S_CTS_ADMI = pd.merge(df_usr02, df_ust04_prof_S_CTS_ADMI, how='left', left_on=['BNAME [C(00024)]'],
                                  right_on='BNAME')
        df_usr02_S_CTS_ADMI = df_usr02_S_CTS_ADMI[df_usr02_S_CTS_ADMI['BNAME'].notna()]
        # print(df_ust12_S_CTS_ADMI)

        # print(df_usr02_S_TCODE['BNAME [C(00024)]'])
        # print(df_usr02_S_TRANSPRT['BNAME [C(00024)]'])
        # print(df_usr02_S_CTS_ADMI)
        # print(df_usr02_S_TCODE,df_usr02_S_TRANSPRT['BNAME [C(00024)]'],df_usr02_S_CTS_ADMI['BNAME [C(00024)]'])
        df_usr02_a=pd.merge(df_usr02_S_TCODE, df_usr02_S_TRANSPRT['BNAME [C(00024)]'], on=['BNAME [C(00024)]'])
        # print(df_usr02_a.columns)
        df_usr02_a = pd.merge(df_usr02_a, df_usr02_S_CTS_ADMI['BNAME [C(00024)]'], on=['BNAME [C(00024)]'])
        # print(df_usr02_a)

        # print(df_usr02_PCFG.shape)

        df_usr02_a = df_usr02_a[
            (df_usr02_a['USTYP [C(00002)]'] == 'A') & (df_usr02_a['UFLAG [b(00001)]'] != 64) & (
                        (df_usr02_a['GLTGB [D(00016)]'] == 0) | (
                            df_usr02_a['GLTGB [D(00016)]'] >= int(self.extract_date)))]
        df_usr02_debug = df_usr02_a[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                         'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
                                         'BNAME [C(00024)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr21 = pd.read_csv(path + '/.USR21.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_a_persno = pd.merge(df_usr02_debug, df_usr21, how='left', left_on=['BNAME [C(00024)]'],
                                         right_on=['bname [C(00024)]'])
        df_adrp = pd.read_csv(path + '/.adrp.CSV', quoting=3, low_memory=False, delimiter='`')
        df_usr02_persno_adrp = pd.merge(df_usr02_a_persno, df_adrp, how='left', left_on=['persnumber [C(00020)]'],
                                        right_on=['persnumber [C(00020)]'])
        df_usr02_persno_adrp = df_usr02_persno_adrp[['CLASS [C(00024)]', 'USTYP [C(00002)]', 'GLTGB [D(00016)]',
                                                     'GLTGV [D(00016)]', 'UFLAG [b(00001)]', 'MANDT [C(00006)]',
                                                     'BNAME [C(00024)]', 'name_first [C(00080)]',
                                                     'name_last [C(00080)]', 'TRDAT [D(00016)]', 'LTIME']]
        df_usr02_persno_adrp.to_excel(excel_writer=self.writer, sheet_name='sap13', index=False,
                                      header=['用户组', '账号类型', '账号有效期至', '有效期自', '锁定状态', '客户名称', '用户', '名', '姓', '最后登录日期',
                                              '最后登录时间'])
        self.writer.save()
        # print( df_prof)
        # print(df_usr02_PCFG.shape)


if __name__ == "__main__":
    sap_process = sap_process_demo('.')
    sap_process.sap18('.')
    sap_process.sap26('.')
    sap_process.sap06('.')
    sap_process.sap06b('.')
    sap_process.sap13('.')
    # sap_process.sap17a('.')
    sap_process.sap05('.')

'''

            f.write(process_str)
            f.close()
            os.system("python ./{}/main.py".format(File.name.split(".rar")[0]))
            print("处理完毕")
            print("开始删除无用的文件")
            files_to_delete = os.listdir()
            for file in files_to_delete:
                if file.startswith(".") and file.endswith("csv"):
                    os.remove(file)

            print("开始下载结果! ")
            file_path = os.path.realpath("SAP_PROCESS_RESULT.xlsx")
            df = pd.read_excel(file_path, sheet_name="sap18")
            usergroup = df["用户组"].tolist()
            account = df["账号类型"].tolist()
            time_ends = df["账号有效期至"].tolist()
            time_starts = df["有效期自"].tolist()
            status = df["锁定状态"].tolist()
            username = df["客户名称"].tolist()
            user = df["用户"].tolist()
            first_name = df["名"].tolist()
            last_name = df["姓"].tolist()
            last_login_date = df["最后登录日期"].tolist()
            last_login_time = df["最后登录时间"].tolist()

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment;filename=' + 'results.xls'
            ws = xlwt.Workbook(encoding='utf-8')

            w = ws.add_sheet('sheet1')
            w.write(0, 0, u'用户组')
            w.write(0, 1, u'账号类型')
            w.write(0, 2, u'账号有效期至')
            w.write(0, 3, u'有效期自')
            w.write(0, 4, u'锁定状态')
            w.write(0, 5, u'客户名称')
            w.write(0, 6, u'用户')
            w.write(0, 7, u'名')
            w.write(0, 8, u'姓')
            w.write(0, 9, u'最后登陆日期')
            w.write(0, 10, u'最后登陆时间')
            excel_row = 1

            for i in range(len(usergroup)):
                w.write(excel_row, 0, usergroup[i])
                w.write(excel_row, 1, account[i])
                w.write(excel_row, 2, time_ends[i])
                w.write(excel_row, 3, time_starts[i])
                w.write(excel_row, 4, status[i])
                w.write(excel_row, 5, username[i])
                w.write(excel_row, 6, user[i])
                w.write(excel_row, 7, first_name[i])
                w.write(excel_row, 8, last_name[i])
                w.write(excel_row, 9, last_login_date[i])
                w.write(excel_row, 10, last_login_time[i])
                excel_row += 1

            output = BytesIO()
            ws.save(output)
            output.seek(0)
            response.write(output.getvalue())
            return response

    return render(request, 'single_choice.html')



def choose_encode(request):
    pass
    return render(request, 'single_choice.html')



def register(request):
    if request.session.get('is_login', None):
        # 登录状态不允许注册。你可以修改这条原则！
        return redirect("/encode/")
    if request.method == "POST":
        register_form = RegisterForm(request.POST)
        message = "请检查填写的内容！"
        if register_form.is_valid():  # 获取数据
            username = register_form.cleaned_data['username']
            password1 = register_form.cleaned_data['password1']
            password2 = register_form.cleaned_data['password2']
            email = register_form.cleaned_data['email']
            if password1 != password2:  # 判断两次密码是否相同
                message = "两次输入的密码不同！"
                return render(request, 'register.html', locals())
            else:
                same_name_user = models.User.objects.filter(name=username)
                if same_name_user:  # 用户名唯一
                    message = '用户已经存在，请重新选择用户名！'
                    return render(request, 'register.html', locals())
                same_email_user = models.User.objects.filter(email=email)
                if same_email_user:  # 邮箱地址唯一
                    message = '该邮箱地址已被注册，请使用别的邮箱！'
                    return render(request, 'register.html', locals())

                # 当一切都OK的情况下，创建新用户
                new_user = models.User.objects.create()
                new_user.name = username
                new_user.password = hash_code(password1)
                new_user.email = email
                new_user.save()
                return redirect('/')  # 自动跳转到登录页面
    register_form = RegisterForm()
    return render(request, 'register.html', locals())


def login(request):
    if request.method == "POST":
        login_form = UserForm(request.POST)
        message = "请检查填写的内容！"
        if login_form.is_valid():
            username = login_form.cleaned_data['username']
            password = login_form.cleaned_data['password']
            try:
                user = models.User.objects.get(name=username)
                if user.password == hash_code(password):
                    request.session['is_login'] = True
                    request.session['user_id'] = user.id
                    request.session['user_name'] = user.name
                    return redirect('/encode/')
                else:
                    message = "密码不正确！"
            except:
                message = "用户不存在！"
        return render(request, 'login.html', locals())

    login_form = UserForm()
    return render(request, 'login.html', locals())


def logout(request):
    if not request.session.get('is_login', None):
        # 如果本来就未登录，也就没有登出一说
        return redirect("/")
    request.session.flush()
    # 或者使用下面的方法
    # del request.session['is_login']
    # del request.session['user_id']
    # del request.session['user_name']
    return redirect('/')
